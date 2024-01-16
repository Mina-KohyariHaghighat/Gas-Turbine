import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side

import input_output


def exhaust_flue_profile(filepath='Fuel/input/input.txt'):

    # calculate volume flow rate of flue

    (projectName,
            gtType,
            temp,
            press,
            rh,
            fuelType,
            LHV,
            volFlowRate,
            massFlowRate,
            exhaustTemp,
            gtLoad) = input_output.txt_exhaust_input(filepath='Fuel/input/input.txt')

    # volMassflow = input_output.txt_exhaust_input.volflowrate
        # gasturbine.performance.gasmass / gasturbine.turbine.outlet_conditions[9]
    volMassflowCorr = int((volFlowRate - 800) / 12.5)+1


    coordinate = pd.read_excel(r'Fuel/databases/Exh-Vel-Profile.csv', sheet_name='X-Y')
    volMassflowRange = pd.read_excel(r'Fuel/databases/Exh-Vel-Profile.csv', sheet_name='massflow')

    # write output file
    par_names = ['X', 'Y', 'X-Velocity', 'Y-Velocity', 'Z-Velocity', 'turb-kinetic-energy', 'turb-diss-rate']
    dest_filename = 'Profile.xlsx'
    wb = Workbook()
    thin = Side(border_style="thin", color="000000")

    for sheet_name in ['Exhaust Velocity']:
        print(' ')
        print(10 * ' ', '{:^20}'.format(sheet_name + ' Profile'), 10 * '')
        row = 1
        ws = wb.create_sheet(sheet_name)

        AmbTemp = round(temp - 273.2, 1)
        lhv = round((lhv) * 1000,0)

        cell_value = ['Exhaust Velocity Vector Profile provided at TUGA',
                        'Project:' +' '+ str(projectName)+ ',' + ' '+ 'Gt. Version:'+' ' +str(gtType),
                        'Tamb:' +' ' + str(AmbTemp)+' '+'°C' + ','+' '+ 'RH:'+ str(rh*100)+' '+'%'\
                        + ','+ ' '+'Pamb:' + str( press/100) + ' ' +'mbar' + ','+ ' '+ 'Gt.Load:'+str( gtLoad)+' '+'%',
                        'Fuel Type:' + ' ' + 'Fuel' + str(fueltype) + ',' + ' ' + 'LHV:' + ' ' + str(lhv) + ' ' + 'KJ/kg']


        for i in range(0, 4):
            c = ws.cell(row + i, 1, cell_value[i ])
            c.alignment = Alignment(horizontal="left", vertical="center")
            if i==0:
                c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=7)
                c.alignment = Alignment(horizontal="center", vertical="center")

            elif i==1 :
                c.border = Border(left=thin, right=thin)
                ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=4)
            elif i == 2:
                c.border = Border(left=thin, right=thin)
                ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=4)
            elif i == 3:
                c.border = Border(left=thin, right=thin, bottom=thin)
                ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=4)
            c.fill = PatternFill("solid", fgColor='FFEBCD')

        cell_value=['Exhaust Mass Flow Rate (Kg/s):',\
                    'Turbine Outlet Temperature (°C):',\
                    'Exhaust Density (Kg/m³):']
        for i in range(1,4):
            c = ws.cell(row + i, 5, cell_value[i-1])
            if i==3:
                c.border = Border( bottom=thin)

            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row + i, start_column=5, end_row=row + i, end_column=6)
            c.fill = PatternFill("solid", fgColor='FFEBCD')

        cell_value=[massFlowRate,exhaustTemp]
        for i in range(1,4):
            c = ws.cell(row + i, 7, cell_value[i-1])
            c.fill = PatternFill("solid", fgColor='FFEBCD')

            if i==3:
                c.border = Border(right=thin, bottom=thin)
            else:
                c.border = Border(right=thin)
            if i==1:
                c.number_format = '0.00'
            if i == 2:
                    c.number_format = '0.0'
            if i == 3:
                c.number_format = '0.0000'
            c.alignment = Alignment(horizontal="left", vertical="center")

        for i in range(7):
            c = ws.cell(5, i + 1, par_names[i])
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor='FF9912')

        for i in range(0, 3091):
            for j in range(2):
                c = ws.cell(i + 6, j + 1, coordinate.iat[i, j])
                c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                c.number_format = '0.0000'
                c.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 14
            if col==6:
                ws.column_dimensions[get_column_letter(col)].width = 20
            ws.alignment = Alignment(horizontal="center", vertical="center")

        '''
        interpolation between to volume mass flow rate
        '''
        k = volMassflowCorr

        if volFlowRate<= 800:
            k=0
        elif volFlowRate >=1350:
            k=44

        volMassflowB = pd.read_excel(r'Fuel/databases/Exh-Vel-Profile.csv', sheet_name=k + 1)
        volMassflowA = pd.read_excel(r'Fuel/databases/Exh-Vel-Profile.csv', sheet_name=k + 2)

        for i in range(0, 3091):
            for j in range(5):
                volMassflow_int = np.array([volMassflowRange.iat[k-1, 1], volMassflowRange.iat[k, 1]])
                vel_int = np.array([volMassflowB.iat[i, j], volMassflowA.iat[i, j]])
                vel_poly = np.polyfit(volMassflow_int, vel_int, 1)
                velocityValue = np.poly1d(vel_poly)
                c = ws.cell(i + 6, j + 3, velocityValue(volFlowRate))
                c.number_format = '0.000'
                c.alignment = Alignment(horizontal='center', textRotation=0)

                c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if fueltype== 'Gas':
                    fueltype='FG'
                elif fueltype== 'Oil':
                    fueltype = 'FO'

        wb.save(filename=str(projectName)+'_'+str(temp)+'°C'+'_'+str(fueltype)+'_'+gtLoad+'_'+dest_filename)
